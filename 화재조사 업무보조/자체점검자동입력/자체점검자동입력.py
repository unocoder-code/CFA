import openpyxl
from openpyxl import *
import time
from tkinter import *
import tkinter.ttk as ttk
import datetime
from tkinter import font
import tkinter.messagebox as msgbox
import threading

#엑셀 마지막 행의 번호
작동기능점검_대상수 = 1517
종합정밀점검_대상수 = 429

print('시작')
# 작동기능점검sheet = 'none'
# 종합정밀점검sheet = 'none'

def 엑셀실행():
    global load_wb
    global 작동기능점검sheet
    global 종합정밀점검sheet
    global photo2
    load_wb = load_workbook(r'\\107.119.40.106\검사반폴더\★2.자체점검대장\0.2021년 자체점검결과보고서 접수대장\2021년 자체점검 대장(작동,종합).xlsx')
    작동기능점검sheet = load_wb['작동기능점검']
    종합정밀점검sheet = load_wb['종합정밀점검']

    photo2 = PhotoImage(file="excel.png")
    mainphoto.config(image=photo2)

def 주소체크1(address):
    row = 0
    작동or종합 = '작동'
    key = address.split()
    #작동에서 찾아보기
    for i in range(8,작동기능점검_대상수+1):
        try:
            주소 = 작동기능점검sheet.cell(i,10).value
            if key[0] in 주소 and key[1] in 주소:
                row = i
                break
        except:
            pass

    #작동에 없을시 -> 종합에서 찾아보기
    if row == 0:
        작동or종합 = '종합'
        for i in range(8,종합정밀점검_대상수+1):
            try:
                주소 = 종합정밀점검sheet.cell(i,10).value
                if key[0] in 주소 and key[1] in 주소:
                    row = i
                    break
            except:
                pass
    return [row,작동or종합]

def 이름체크1(name):
    row = 0
    작동or종합 = 'None'
    #작동에서 찾아보기
    for i in range(8,작동기능점검_대상수+1):
        try:
            이름 = 작동기능점검sheet.cell(i,9).value
            if name in 이름:
                작동or종합 = '작동'
                row = i
                break
        except:
            pass

    #작동에 없을시 -> 종합에서 찾아보기
    if row == 0:
        
        for i in range(8,종합정밀점검_대상수+1):
            try:
                이름 = 종합정밀점검sheet.cell(i,9).value
                if name in 이름:
                    작동or종합 = '종합'
                    row = i
                    break
            except:
                pass
    return [row,작동or종합]

def 대상_검색하기():
    global target
    global statenum
    statenum = 0
    target = [0,'None']
    address1 = 검색_주소.get()
    name1 = 검색_이름.get()

    if len(address1) < 2:
        address1 = 'XXX'
    if len(name1) < 2:
        name1 = 'XXX'

    target = 주소체크1(address1)
    if target[0]==0:
        target = 이름체크1(name1)

    #검색 끝-
    if target[0] == 0:
        결과_이름2.config(text="대상을 찾을 수 없습니다")
        결과_주소2.config(text="대상을 찾을 수 없습니다")
    else:
        if target[1] == '작동':
            결과_이름2.config(text=작동기능점검sheet.cell(target[0],9).value)
            결과_주소2.config(text=작동기능점검sheet.cell(target[0],10).value)
        elif target[1] == '종합':
            결과_이름2.config(text=종합정밀점검sheet.cell(target[0],9).value)
            결과_주소2.config(text=종합정밀점검sheet.cell(target[0],10).value)


def 정보등록():
    if target[0] < 8:
        msgbox.showwarning("에러", "대상을 검색하세요")
        return

    data = [점검자.get(),점검기간.get(),결과보고.get(),양호일까조치명령일까.get()]
    print(target,data)
    if target[1] == '작동':
        if 작동일까종합일까.get() == '종합':
            return
        for i in range(4):
            작동기능점검sheet.cell(target[0], i+18).value = data[i]

    elif target[1] == '종합':
        작동or종합결과 = 작동일까종합일까.get()
        if 작동or종합결과 == '작동':
            종합정밀점검sheet.cell(target[0], 25).value = '제출완료'
            for i in range(4):
                종합정밀점검sheet.cell(target[0], i+26).value = data[i]
        elif 작동or종합결과 == '종합':
            종합정밀점검sheet.cell(target[0], 19).value = '제출완료'
            for i in range(4):
                종합정밀점검sheet.cell(target[0], i+20).value = data[i]

    msgbox.showinfo("알림","정보 등록 완료")

def 정보등록1():
    정보등록Thread= threading.Thread(target=정보등록)
    정보등록Thread.start()

def 엑셀저장하기():
    load_wb.save(r'\\107.119.40.106\검사반폴더\★2.자체점검대장\0.2021년 자체점검결과보고서 접수대장\2021년 자체점검 대장(작동,종합).xlsx')
    msgbox.showinfo("알림","엑셀 파일 저장 완료")

def gui_program():
    global 결과_이름2
    global 결과_주소2
    global 검색_이름
    global 검색_주소
    global 작동일까종합일까
    global 양호일까조치명령일까
    global 점검자
    global 점검기간
    global 결과보고
    global mainphoto

    root = Tk()
    root.title("자체점검 엑셀")
    root.geometry("450x400")
    
    row_num = 0
    blank1 = Label(root,text="     ")
    blank1.grid(row=row_num, column=0)

    photo = PhotoImage(file="Ready.png")
    mainphoto = Label(root, image=photo)
    mainphoto.grid(row=row_num, column=1,rowspan=2, sticky=E)

    프로그램제목 = Label(root,text="자체점검자동입력프로그램",font=font.Font(family="맑은 고딕",size = 13,weight='bold'))
    프로그램제목.grid(row=row_num, column=2,rowspan=2, columnspan=2, sticky=W)
    
    row_num+=1
    
    blank1 = Label(root,text="     ")
    blank1.grid(row=row_num, column=0)

    row_num+=1

    이름label = Label(root,text="대상 이름",font=font.Font(family="맑은 고딕",size = 9,weight='bold'))
    이름label.grid(row=row_num, column=1, sticky=E)

    검색_이름 = Entry(root, width=20)
    검색_이름.grid(row=row_num, column=2, sticky=W)
    #검색_이름.insert(0,"대상 이름을 입력하세요")

    대상검색버튼 = Button(root, width=6,height=3, text="검색", command=대상_검색하기)
    대상검색버튼.grid(row=row_num, column=3, rowspan= 2, sticky=W)
    row_num+=1

    주소label = Label(root,text="대상 주소",font=font.Font(family="맑은 고딕",size = 9,weight='bold'))
    주소label.grid(row=row_num, column=1, sticky=E)

    검색_주소 = Entry(root, width=20)
    검색_주소.grid(row=row_num, column=2, sticky=W)
    #주소_이름.insert(0,"대상 주소를 입력하세요")

    row_num+=1

    결과label = Label(root,text="#검색 결과#",font=font.Font(family="맑은 고딕",size = 11,weight='bold'))
    결과label.grid(row=row_num, column=2, sticky=W)
    
    row_num+=1
    

    결과_이름 = Label(root,text="이름 :   ",font=font.Font(family="맑은 고딕",size = 9,weight='bold'))
    결과_이름.grid(row=row_num, column=1, sticky=E)

    결과_이름2 = Label(root,text="None")
    결과_이름2.grid(row=row_num, column=2, columnspan=2, sticky=W)
    
    row_num+=1

    결과_주소 = Label(root,text="주소 :   ",font=font.Font(family="맑은 고딕",size = 9,weight='bold'))
    결과_주소.grid(row=row_num, column=1, sticky=E)

    결과_주소2 = Label(root,text="None")
    결과_주소2.grid(row=row_num, column=2, columnspan=2, sticky=W)
    
    row_num+=1
    
    blank1 = Label(root,text="--------------------------------------------------------------------------------------")
    blank1.grid(row=row_num, column=0, columnspan=4)
    
    row_num+=1
    작동종합label = Label(root,text="    작동or조치",font=font.Font(family="맑은 고딕",size = 9,weight='bold'))
    작동종합label.grid(row=row_num, column=1, sticky=W)
    작동일까종합일까 = StringVar()
    작동이지롱 = Radiobutton(root, text= "작동",value = '작동' , variable=작동일까종합일까)
    작동이지롱.select()
    종합이지롱 = Radiobutton(root, text= "종합           ",value = '종합', variable=작동일까종합일까)
    
    작동이지롱.grid(row=row_num, column=2,sticky=W)
    종합이지롱.grid(row=row_num, column=2,sticky=E)

    row_num+=1

    점검자label = Label(root,text="       점검자",font=font.Font(family="맑은 고딕",size = 9,weight='bold'))
    점검자label.grid(row=row_num, column=1, sticky=W)

    점검자_리스트 = ['에이스방재','중앙방재','소방안전관리자']
    점검자 = ttk.Combobox(root, width=17, height=3, values=점검자_리스트)
    점검자.grid(row=row_num, column=2,sticky=W)
    점검자.set("중앙방재") #최초 목록 제목 설정

    결과입력버튼 = Button(root, width=6,height=3, text="등록", command=정보등록1)
    결과입력버튼.grid(row=row_num, column=3, rowspan=3,sticky=W)

    row_num+=1

    점검기간label = Label(root,text="      점검기간",font=font.Font(family="맑은 고딕",size = 9,weight='bold'))
    점검기간label.grid(row=row_num, column=1, sticky=W)

    점검기간_리스트 = [(now-datetime.timedelta(i)).strftime('%Y.%m.%d') for i in range (8)]
    점검기간 = ttk.Combobox(root, width=17, height=5, values=점검기간_리스트)
    점검기간.grid(row=row_num, column=2, sticky=W)
    점검기간.set(now.strftime('%Y.%m.%d')) #최초 목록 제목 설정

    row_num+=1

    결과보고label = Label(root,text="      결과보고",font=font.Font(family="맑은 고딕",size = 9,weight='bold'))
    결과보고label.grid(row=row_num, column=1, sticky=W)

    결과보고_리스트 = [(now-datetime.timedelta(i)).strftime('%Y.%m.%d') for i in range (3)]
    결과보고 = ttk.Combobox(root, width=17, height=3, values=결과보고_리스트)
    결과보고.grid(row=row_num, column=2, sticky=W)
    결과보고.set(now.strftime('%Y.%m.%d')) #최초 목록 제목 설정

    row_num+=1

    양호조치label = Label(root,text="    양호or조치",font=font.Font(family="맑은 고딕",size = 9,weight='bold'))
    양호조치label.grid(row=row_num, column=1, sticky=W)

    양호일까조치명령일까 = StringVar()
    양호지롱 = Radiobutton(root, text= "양호",value = '양호' , variable=양호일까조치명령일까)
    양호지롱.select()
    조치명령이지롱 = Radiobutton(root, text= "조치명령     ",value = '조치명령', variable=양호일까조치명령일까)
    
    양호지롱.grid(row=row_num, column=2,sticky=W)
    조치명령이지롱.grid(row=row_num, column=2,sticky=E)

    row_num+=1
    
    blank1 = Label(root,text="  ")
    blank1.grid(row=row_num, column=0, columnspan=4)
    row_num+=1

    파일저장버튼 = Button(root, fg="yellow", bg="green", width=10,height=2, text="엑셀 저장",font=font.Font(family="맑은 고딕",size = 12,weight='bold') , command=엑셀저장하기)
    파일저장버튼.grid(row=row_num, column=1)

    재시작버튼 = Button(root, bg="orange", width=10,height=2, text="X",font=font.Font(family="맑은 고딕",size = 12,weight='bold')) #, command=root.quit())
    재시작버튼.grid(row=row_num, column=2)

    종료버튼 = Button(root, fg="red", bg="yellow", width=10,height=2, text="종료",font=font.Font(family="맑은 고딕",size = 12,weight='bold') , command=root.destroy)
    종료버튼.grid(row=row_num, column=3)

    root.mainloop()

if __name__ == '__main__':
    now = datetime.datetime.today()
    target = [0,'None']

    엑셀실행 = threading.Thread(target=엑셀실행)
    엑셀실행.start()

    gui_program()