from json import load, loads
from re import search
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import *
import time
from selenium.webdriver.remote.utils import format_json
from tkinter import *
import threading

#from 건축분야함수 import *
#from 소방분야함수 import *

def check_loading():
    try:
        html_raw = driver.page_source
        html = BeautifulSoup(html_raw, 'html.parser')
        loadingcode = html.find(id="fsseProgress")
        try:
            load_attrs = loadingcode.attrs
            if 'in' in load_attrs['class']:
                print ('*', end='')
                time.sleep(1)
                check_loading()
            else:
                time.sleep(1)
                print('\nEnd Check loading')
        except:
            print("Error")
            time.sleep(100)
            pass
    except:
        check_loading()
        pass

def 점검자서명1():
    try:
        click('//*[@id="tab9"]/div/div/table/tbody/tr[1]/th[2]/button')
        click('//*[@id="tab9"]/div/div/table/tbody/tr[1]/th[2]/button')
        time.sleep(0.1)
        click('//*[@id="tab9_iptr_type-1"]/option[4]')
        click('//*[@id="tab9_iptr_type-2"]/option[4]')
        time.sleep(0.1)
        sendkey('//*[@id="tab9_iptn-1"]','이대진')
        sendkey('//*[@id="tab9_iptn-2"]','송병진')
        time.sleep(0.1)
    except:
        pass

def 점검자서명2():
    try:
        click('//*[@id="tab9"]/div/div/table/tbody/tr[1]/th[2]/button')
        click('//*[@id="tab9"]/div/div/table/tbody/tr[1]/th[2]/button')
        time.sleep(0.1)
        click('//*[@id="tab9_iptr_type-1"]/option[4]')
        click('//*[@id="tab9_iptr_type-2"]/option[4]')
        time.sleep(0.1)
        sendkey('//*[@id="tab9_iptn-1"]','신학균')
        sendkey('//*[@id="tab9_iptn-2"]','정조영')
        time.sleep(0.1)
    except:
        pass

def 건축물개요_입력1():
    개요Thread= threading.Thread(target=건축물개요_입력)
    개요Thread.start()

def 건축물개요_입력():
    try:
        load_wb = load_workbook('C:/Users/win10/Desktop/대상처/{}/건축물개요_이름.xlsx'.format(building_name), data_only=True)
        #load_wb = load_workbook('C:/Users/win10/Desktop/자동화 프로그램/화재안전특별조사v2/건축물개요_이름.xlsx', data_only=True)
        건축물개요정보 = []
        주요시설 = []

        load_ws = load_wb['건축물개요']
        for i in range (2,18):
            건축물개요정보.append(load_ws.cell(i,2).value)
        for i in range (2,20):
            주요시설.append(load_ws.cell(i,5).value)

        click('//*[@id="tab_title1"]')
        time.sleep(1)
        if 건축물개요정보[0] != 0:
            driver.find_element_by_xpath('//*[@id="tab1_fon"]').send_keys(Keys.CONTROL, 'a')
            driver.find_element_by_xpath('//*[@id="tab1_fon"]').send_keys(Keys.BACKSPACE)
            sendkey('//*[@id="tab1_fon"]',건축물개요정보[0])
            driver.find_element_by_xpath('//*[@id="etc_addr"]').send_keys(Keys.CONTROL, 'a')
            driver.find_element_by_xpath('//*[@id="etc_addr"]').send_keys(Keys.BACKSPACE)
            sendkey('//*[@id="etc_addr"]',건축물개요정보[0])
        if 건축물개요정보[1] != 0:
            click('//*[@id="tab1_main_prps"]/option[{}]'.format(건축물개요정보[1]))
        if 건축물개요정보[2] != 0:
            click('//*[@id="tab1_main_strc"]/option[{}]'.format(건축물개요정보[2]))
        if 건축물개요정보[3] != 0:
            click('//*[@id="tab1_rf_strc"]/option[{}]'.format(건축물개요정보[3]))
        if 건축물개요정보[4] != 0:
            driver.find_element_by_xpath('//*[@id="tab1_prms_de"]').send_keys(Keys.CONTROL, 'a')
            driver.find_element_by_xpath('//*[@id="tab1_prms_de"]').send_keys(Keys.BACKSPACE)
            sendkey('//*[@id="tab1_prms_de"]',건축물개요정보[4])
        if 건축물개요정보[5] != 0:
            driver.find_element_by_xpath('//*[@id="tab1_use_dt"]').send_keys(Keys.CONTROL, 'a')
            driver.find_element_by_xpath('//*[@id="tab1_use_dt"]').send_keys(Keys.BACKSPACE)
            sendkey('//*[@id="tab1_use_dt"]',건축물개요정보[5])
        if 건축물개요정보[6] != 0:
            click('//*[@id="tab1_insc_cpnm"]/option[{}]'.format(건축물개요정보[6]))
        if 건축물개요정보[7] != 0:
            driver.find_element_by_xpath('//*[@id="tab1_pssc_amt"]').send_keys(Keys.CONTROL, 'a')
            driver.find_element_by_xpath('//*[@id="tab1_pssc_amt"]').send_keys(Keys.BACKSPACE)
            sendkey('//*[@id="tab1_pssc_amt"]',건축물개요정보[7])
        if 건축물개요정보[8] != 0:
            driver.find_element_by_xpath('//*[@id="tab1_scat_amt"]').send_keys(Keys.CONTROL, 'a')
            driver.find_element_by_xpath('//*[@id="tab1_scat_amt"]').send_keys(Keys.BACKSPACE)
            sendkey('//*[@id="tab1_scat_amt"]',건축물개요정보[8])
        if 건축물개요정보[9] != 0:
            driver.find_element_by_xpath('//*[@id="tab1_insc_bgin_dt"]').send_keys(Keys.CONTROL, 'a')
            driver.find_element_by_xpath('//*[@id="tab1_insc_bgin_dt"]').send_keys(Keys.BACKSPACE)
            sendkey('//*[@id="tab1_insc_bgin_dt"]',건축물개요정보[9])
        if 건축물개요정보[10] != 0:
            driver.find_element_by_xpath('//*[@id="tab1_insc_end_dt"]').send_keys(Keys.CONTROL, 'a')
            driver.find_element_by_xpath('//*[@id="tab1_insc_end_dt"]').send_keys(Keys.BACKSPACE)
            sendkey('//*[@id="tab1_insc_end_dt"]',건축물개요정보[10])

        if 건축물개요정보[12] != 0:
            driver.find_element_by_xpath('//*[@id="tab1_bldf_ar"]').send_keys(Keys.CONTROL, 'a')
            driver.find_element_by_xpath('//*[@id="tab1_bldf_ar"]').send_keys(Keys.BACKSPACE)
            sendkey('//*[@id="tab1_bldf_ar"]',건축물개요정보[12])

        if 건축물개요정보[13] != 0:
            driver.find_element_by_xpath('//*[@id="tab1_bdng_ar"]').send_keys(Keys.CONTROL, 'a')
            driver.find_element_by_xpath('//*[@id="tab1_bdng_ar"]').send_keys(Keys.BACKSPACE)
            sendkey('//*[@id="tab1_bdng_ar"]',건축물개요정보[13])

        if 건축물개요정보[14] != 0:
            driver.find_element_by_xpath('//*[@id="tab1_grnd_fl"]').send_keys(Keys.CONTROL, 'a')
            driver.find_element_by_xpath('//*[@id="tab1_grnd_fl"]').send_keys(Keys.BACKSPACE)
            sendkey('//*[@id="tab1_grnd_fl"]',건축물개요정보[14])

        if 건축물개요정보[15] != 0:
            driver.find_element_by_xpath('//*[@id="tab1_udgd_fl"]').send_keys(Keys.CONTROL, 'a')
            driver.find_element_by_xpath('//*[@id="tab1_udgd_fl"]').send_keys(Keys.BACKSPACE)
            sendkey('//*[@id="tab1_udgd_fl"]',건축물개요정보[15])

        for i in range(1,10):
            if 주요시설[i-1] == 'o':
                click('//*[@id="chk_tab1_main_fcty_VN001004000{}"]'.format(i))
        for i in range(10,19):
            if 주요시설[i-1] == 'o':
                click('//*[@id="chk_tab1_main_fcty_VN00100400{}"]'.format(i))

        #날짜 선택
        click('//*[@id="tbl_buld"]/tbody/tr[4]/td/button/img')
        time.sleep(1)
        click('//*[@id="ui-datepicker-div"]/table/tbody/tr[{}]/td[{}]/a'.format(int((건축물개요정보[11]+3)/7)+1,((건축물개요정보[11]+3)%7)+1))
        time.sleep(1)

        click('//*[@id="b_Insert"]')
        try:
            click('//*[@id="btn_ok_c"]')
        except:
            time.sleep(1)
            click('//*[@id="btn_ok_c"]')
        time.sleep(1)
        try:
            click('//*[@id="btn_ok_a"]')
        except:
            time.sleep(1)
            click('//*[@id="btn_ok_a"]')
    except:
        pass

def 건축분야1():
    건축분야Thread= threading.Thread(target=건축분야)
    건축분야Thread.start()

def 건축분야():
    try:
        load_wb = load_workbook('C:/Users/win10/Desktop/대상처/{}/건축분야형식_이름.xlsx'.format(building_name))
        click('//*[@id="tab_title3"]') #건축분야
        time.sleep(1)
        click('//*[@id="btn_show_A"]') #전체펼치기
        time.sleep(1)
        방화구획클릭(load_wb)
        피난시설클릭(load_wb)
        방화문클릭(load_wb)
        방화셔터클릭(load_wb)
        마감재클릭(load_wb)
        주차장클릭(load_wb)
        click('//*[@id="b_Insert"]')
        try:
            click('//*[@id="btn_ok_c"]')
        except:
            time.sleep(1)
            click('//*[@id="btn_ok_c"]')
        time.sleep(1)
        try:
            click('//*[@id="btn_ok_a"]')
        except:
            time.sleep(1)
            click('//*[@id="btn_ok_a"]')

        load_wb.close()
    except:
        pass

def click_GNB1(x,i):
    if x != 'N':
        click('//*[@id="dtck{}_A000{}"]'.format(x,i))
        if x == 'B':
            click('//*[@id="itm_A000{}APL0001"]'.format(i))

def click_GNB(x,i):
    if x != 'N':
        click('//*[@id="dtck{}_A00{}"]'.format(x,i))
        if x == 'B':
            click('//*[@id="itm_A00{}APL0001"]'.format(i))

#G B N
def 방화구획클릭(load_wb):
    방화구획 = [2,'N','N','N','N','N','N','N']

    load_ws = load_wb['1방화구획']
    for i in range (1,9):
        방화구획[i-1] = load_ws.cell(i,2).value
    
    a = 0
    click('//*[@id="itm_A0001APL000{}"]'.format(방화구획[a]))
    a+=1
    for i in range(2,7):
        click_GNB1(방화구획[a],i)
        a+=1
    for i in range(8,10):
        click_GNB1(방화구획[a],i)
        a+=1

def 피난시설클릭(load_wb): # 12~19
    #직통계단
    피난시설 = ['N','N','N','N','N','N','N','N','N','N','N','N','N']

    load_ws = load_wb['2피난시설']
    for i in range (1,11):
        피난시설[i-1] = load_ws.cell(i,2).value

    a = 0
    for i in range(12,20):
        click_GNB(피난시설[a],i)
        a+=1

    if 피난시설[8] != 'N':
        a = 0
        for i in range(21,29):
            click_GNB(피난시설[a],i)
            a+=1
        click('//*[@id="dtck{}_A0029"]'.format(피난시설[8]))
        if 피난시설[8] == 'B':
            click('//*[@id="itm_A0029APL0001"]')

    if 피난시설[9] != 'N':
        a = 0
        for i in range(32,40):
            click_GNB(피난시설[a],i)
            a+=1
        click('//*[@id="dtck{}_A0040"]'.format(피난시설[9]))
        if 피난시설[9] == 'B':
            click('//*[@id="itm_A0040APL0001"]')

    if 피난시설[10] != 'N':
        click_GNB(피난시설[a],i)

        a = 5
        for i in range(48,51):
            click_GNB(피난시설[a],i)
            a+=1

        click('//*[@id="dtck{}_A0051"]'.format(피난시설[10]))
        if 피난시설[a] == 'B':
            click('//*[@id="itm_A0051APL0001"]')

    if 피난시설[11] != 'N':
        click_GNB(피난시설[a],61)
    
    if 피난시설[12] != 'N':
        click_GNB(피난시설[a],71)

def 방화문클릭(load_wb):
    방화문 = [2,'N','N','N','N','N','N','N']

    load_ws = load_wb['3방화문']
    for i in range (1,9):
        방화문[i-1] = load_ws.cell(i,2).value

    a = 0
    click('//*[@id="itm_A0072APL000{}"]'.format(방화문[a]))
    a+=1
    for i in range(73,80):
        click_GNB(방화문[a],i)
        a+=1

def 방화셔터클릭(load_wb):
    방화셔터 = [2,'N','N','N','N']

    load_ws = load_wb['4방화셔터']
    for i in range (1,6):
        방화셔터[i-1] = load_ws.cell(i,2).value

    a = 0
    click('//*[@id="itm_A0080APL000{}"]'.format(방화셔터[a]))
    a+=1
    for i in range(82,85):
        click_GNB(방화셔터[a],i)
        a+=1

def 마감재클릭(load_wb):
    마감재 = [4,4,4]

    load_ws = load_wb['5마감재']
    for i in range (1,4):
        마감재[i-1] = load_ws.cell(i,2).value

    a = 0
    click('//*[@id="itm_A0087APL000{}"]'.format(마감재[a]))
    a+=1
    click('//*[@id="itm_A0089APL000{}"]'.format(마감재[a]))
    a+=1
    click('//*[@id="itm_A0091APL000{}"]'.format(마감재[a]))

def 주차장클릭(load_wb):
    주차장 = [0,0]

    load_ws = load_wb['주차장']
    for i in range (1,3):
        주차장[i-1] = load_ws.cell(i,2).value

    a = 0
    if 주차장[0] != 0:
        sendkey('//*[@id="dtck_A00{}"]'.format(주차장[0]+96),주차장[1])

def 소방분야1():
    소방Thread = threading.Thread(target = 소방분야)
    소방Thread.start()

def 소방분야():
    try:
        load_wb = load_workbook('C:/Users/win10/Desktop/대상처/{}/소방분야형식_이름.xlsx'.format(building_name))
        click('//*[@id="tab_title4"]') #소방분야
        time.sleep(1)
        click('//*[@id="btn_show_S"]') #전체펼치기
        time.sleep(1)
        자체관리분야클릭(load_wb)
        소화기구클릭(load_wb)
        자동소화장치클릭(load_wb)
        수계소화설비클릭(load_wb)
        옥내외소화전클릭(load_wb)
        간이스프링클러클릭(load_wb)
        가스계설비클릭(load_wb)
        경보설비클릭(load_wb)
        피난설비클릭(load_wb)
        소화용수설비클릭(load_wb)
        재연설비클릭(load_wb)
        연결송수관클릭(load_wb)
        비상콘센트_피난방화클릭(load_wb)
        방염물품_화기취급클릭(load_wb)
        위험물저장클릭(load_wb)
        소방활동관사클릭(load_wb)
        이용자특성클릭(load_wb)
        관리자특성클릭(load_wb)
        click('//*[@id="b_Insert"]')
        try:
            click('//*[@id="btn_ok_c"]')
        except:
            time.sleep(1)
            click('//*[@id="btn_ok_c"]')
        time.sleep(1)
        try:
            click('//*[@id="btn_ok_a"]')
        except:
            time.sleep(1)
            click('//*[@id="btn_ok_a"]')

        load_wb.close()
    except:
        pass

#G B N
def 자체관리분야클릭(load_wb):
    자체관리분야 = [5,'N',0,'N','N','N','N','N',2]

    load_ws = load_wb['1자체관리분야']
    for i in range (1,10):
        자체관리분야[i-1] = load_ws.cell(i,2).value

    a = 0
    click('//*[@id="itm_S0002SPL000{}"]'.format(자체관리분야[a]))
    a+=1
    if 자체관리분야[a] != 'N':
        click('//*[@id="dtck{}_S0003"]'.format(자체관리분야[a]))
        if 자체관리분야[a] == 'B':
                click('//*[@id="itm_S0003SPL0002"]')
    a+=1
    if 자체관리분야[a] != 0:
        click('//*[@id="itm_S0004SPL000{}"]'.format(자체관리분야[a]))
    a+=1
    if 자체관리분야[a] != 'N':
        click('//*[@id="dtck{}_S0005"]'.format(자체관리분야[a]))
        if 자체관리분야[a] == 'B':
                click('//*[@id="itm_S0005SPL0002"]')
    a+=1
    
    for i in range(7,10):
        if 자체관리분야[a] != 'N':
            click('//*[@id="dtck{}_S000{}"]'.format(자체관리분야[a],i))
            if 자체관리분야[a] == 'B':
                click('//*[@id="itm_S000{}SPL0002"]'.format(i))
        a+=1
    if 자체관리분야[a] != 'N':
        click('//*[@id="dtck{}_S0010"]'.format(자체관리분야[a],i))
        if 자체관리분야[a] == 'B':
            click('//*[@id="itm_S0010SPL0002"]')
    a+=1

    click('//*[@id="itm_S0011SPL000{}"]'.format(자체관리분야[a]))
    if 자체관리분야[a] == 1:
            click('//*[@id="itm_S0012SPL0006"]')

def 소화기구클릭(load_wb):
    소화기구 = ['N','N']

    load_ws = load_wb['2소화기구']
    for i in range (1,3):
        소화기구[i-1] = load_ws.cell(i,2).value

    a = 0
    for i in range(14,16):
        if 소화기구[a] != 'N':
            click('//*[@id="dtck{}_S00{}"]'.format(소화기구[a],i))
            if 소화기구[a] == 'B':
                click('//*[@id="itm_S00{}SPL0001"]'.format(i))
        a+=1

def 자동소화장치클릭(load_wb):
    자동소화장치 = ['N','N','N','N']

    load_ws = load_wb['3자동소화장치']
    for i in range (1,5):
        자동소화장치[i-1] = load_ws.cell(i,2).value

    a = 0
    for i in range(17,21):
        if 자동소화장치[a] != 'N':
            click('//*[@id="dtck{}_S00{}"]'.format(자동소화장치[a],i))
            if 자동소화장치[a] == 'B':
                click('//*[@id="itm_S00{}SPL0001"]'.format(i))
        a+=1

def 수계소화설비클릭(load_wb):
    수계소화설비 = [3,3,3,'N','N','N','N','N','N','N','N','N']

    load_ws = load_wb['4수계소화설비']
    for i in range (1,13):
        수계소화설비[i-1] = load_ws.cell(i,2).value

    a = 0
    click('//*[@id="itm_S0023SPL000{}"]'.format(수계소화설비[a]))
    a+=1
    click('//*[@id="itm_S0024SPL000{}"]'.format(수계소화설비[a]))
    a+=1
    click('//*[@id="itm_S0025SPL000{}"]'.format(수계소화설비[a]))
    a+=1

    for i in range(26,32):
        if 수계소화설비[a] != 'N':
            click('//*[@id="dtck{}_S00{}"]'.format(수계소화설비[a],i))
            if 수계소화설비[a] == 'B':
                click('//*[@id="itm_S00{}SPL0001"]'.format(i))
        a+=1
    for i in range(33,36):
        if 수계소화설비[a] != 'N':
            click('//*[@id="dtck{}_S00{}"]'.format(수계소화설비[a],i))
            if 수계소화설비[a] == 'B':
                click('//*[@id="itm_S00{}SPL0001"]'.format(i))
        a+=1

def 옥내외소화전클릭(load_wb): #+포혼합
    옥내외소화전 = ['N','N','N','N','N','N','N']

    load_ws = load_wb['5옥내외소화전']
    for i in range (1,8):
        옥내외소화전[i-1] = load_ws.cell(i,2).value

    a = 0
    for i in range(38,40):
        if 옥내외소화전[a] != 'N':
            click('//*[@id="dtck{}_S00{}"]'.format(옥내외소화전[a],i))
            if 옥내외소화전[a] == 'B':
                click('//*[@id="itm_S00{}SPL0001"]'.format(i))
        a+=1
    if 옥내외소화전[a] != 'N':
        click('//*[@id="dtck{}_S0040"]'.format(옥내외소화전[a]))
        if 옥내외소화전[a] == 'B':
            click('//*[@id="itm_S0040SPL0006"]')
    a+=1
    for i in range(42,44):
        if 옥내외소화전[a] != 'N':
            click('//*[@id="dtck{}_S00{}"]'.format(옥내외소화전[a],i))
            if 옥내외소화전[a] == 'B':
                click('//*[@id="itm_S00{}SPL0001"]'.format(i))
        a+=1
    if 옥내외소화전[a] != 'N':
        click('//*[@id="dtck{}_S0044"]'.format(옥내외소화전[a]))
        if 옥내외소화전[a] == 'B':
            click('//*[@id="itm_S0044SPL0006"]')
    a+=1

    if 옥내외소화전[a] != 'N':
        click('//*[@id="dtck{}_S0046"]'.format(옥내외소화전[a]))
        if 옥내외소화전[a] == 'B':
            click('//*[@id="itm_S0046SPL0001"]')
    a+=1

def 간이스프링클러클릭(load_wb):
    간이스프링클러 = [4,4,4,4,4]
    가보자 = []
    load_ws = load_wb['7간이스프링클러']
    for i in range (1,6):
        간이스프링클러[i-1] = load_ws.cell(i,2).value
    for i in range (7,16):
        가보자.append(load_ws.cell(i,2).value)
    a = 0
    click('//*[@id="itm_S0049SPL000{}"]'.format(간이스프링클러[a]))
    a+=1
    click('//*[@id="itm_S0061SPL000{}"]'.format(간이스프링클러[a]))
    a+=1
    click('//*[@id="itm_S0073SPL000{}"]'.format(간이스프링클러[a]))
    a+=1
    click('//*[@id="itm_S0085SPL000{}"]'.format(간이스프링클러[a]))
    a+=1
    click('//*[@id="itm_S0097SPL000{}"]'.format(간이스프링클러[a]))

    if 가보자[0] == 0:
        pass
    elif 가보자[0] == 1:
        for j in range (51,55):
            click('//*[@id="dtck{}_S00{}"]'.format(가보자[j-50],j))
        for j in range (56,60):
            click('//*[@id="dtck{}_S00{}"]'.format(가보자[j-51],j))
    elif 가보자[0] == 2:
        for j in range (63,67):
            click('//*[@id="dtck{}_S00{}"]'.format(가보자[j-62],j))
        for j in range (68,72):
            click('//*[@id="dtck{}_S00{}"]'.format(가보자[j-63],j))
    elif 가보자[0] == 3:
        for j in range (75,79):
            click('//*[@id="dtck{}_S00{}"]'.format(가보자[j-74],j))
        for j in range (80,84):
            click('//*[@id="dtck{}_S00{}"]'.format(가보자[j-75],j))
    elif 가보자[0] == 4:
        for j in range (87,91):
            click('//*[@id="dtck{}_S00{}"]'.format(가보자[j-86],j))
        for j in range (92,96):
            click('//*[@id="dtck{}_S00{}"]'.format(가보자[j-87],j))
    elif 가보자[0] == 5:
        for j in range (99,103):
            click('//*[@id="dtck{}_S00{}"]'.format(가보자[j-98],j))
        for j in range (104,108):
            click('//*[@id="dtck{}_S00{}"]'.format(가보자[j-99],j))

def 가스계설비클릭(load_wb):
    가스계설비 = ['N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N']
    
    load_ws = load_wb['8가스계설비']
    for i in range (1,37):
        가스계설비[i-1] = load_ws.cell(i,2).value

    a = 0
    for i in range(110,116):
        if 가스계설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(가스계설비[a],i))
            if 가스계설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(117,120):
        if 가스계설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(가스계설비[a],i))
            if 가스계설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(121,127):
        if 가스계설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(가스계설비[a],i))
            if 가스계설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(128,131):
        if 가스계설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(가스계설비[a],i))
            if 가스계설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(132,138):
        if 가스계설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(가스계설비[a],i))
            if 가스계설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(139,142):
        if 가스계설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(가스계설비[a],i))
            if 가스계설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(143,149):
        if 가스계설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(가스계설비[a],i))
            if 가스계설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(150,153):
        if 가스계설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(가스계설비[a],i))
            if 가스계설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1

def 경보설비클릭(load_wb):
    경보설비 = [0,'N','N','N','N','N','N','N','N','N','N']

    load_ws = load_wb['9경보설비']
    for i in range (1,12):
        경보설비[i-1] = load_ws.cell(i,2).value

    a = 0
    if 경보설비[a] != 0:
        click('//*[@id="itm_S0154SPL000{}"]'.format(경보설비[a]))
    a+=1
    for i in range(155,162):
        if 경보설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(경보설비[a],i))
            if 경보설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(163,165):
        if 경보설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(경보설비[a],i))
            if 경보설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    if 경보설비[a] != 'N':
        click('//*[@id="dtck{}_S0166"]'.format(경보설비[a]))
        if 경보설비[a] == 'B':
                click('//*[@id="itm_S00166SPL0001"]'.format(i))

def 피난설비클릭(load_wb): #+포혼합
    피난설비 = ['N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N']
    
    load_ws = load_wb['10피난설비']
    for i in range (1,24):
        피난설비[i-1] = load_ws.cell(i,2).value

    a = 0
    for i in range(169,174): #유도등
        if 피난설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(피난설비[a],i))
            if 피난설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(175,178): #비상조명
        if 피난설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(피난설비[a],i))
            if 피난설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1

    for i in range(180,183): #완강기
        if 피난설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(피난설비[a],i))
            if 피난설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(184,187): #구조대
        if 피난설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(피난설비[a],i))
            if 피난설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(188,191): #미끄럼대
        if 피난설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(피난설비[a],i))
            if 피난설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(192,195): #습강식
        if 피난설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(피난설비[a],i))
            if 피난설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(196,199): #기타
        if 피난설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(피난설비[a],i))
            if 피난설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1

def 소화용수설비클릭(load_wb):
    소화용수설비 = ['N','N','N','N','N']

    load_ws = load_wb['11소화용수설비']
    for i in range (1,6):
        소화용수설비[i-1] = load_ws.cell(i,2).value

    a = 0
    for i in range(200,205):
        if 소화용수설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(소화용수설비[a],i))
            if 소화용수설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1

def 재연설비클릭(load_wb):
    재연설비 = ['N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N','N']
    
    load_ws = load_wb['12재연설비']
    for i in range (1,19):
        재연설비[i-1] = load_ws.cell(i,2).value

    a = 0
    for i in range(207,210):
        if 재연설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(재연설비[a],i))
            if 재연설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(211,214):
        if 재연설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(재연설비[a],i))
            if 재연설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(215,218):
        if 재연설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(재연설비[a],i))
            if 재연설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(219,222):
        if 재연설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(재연설비[a],i))
            if 재연설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(223,226):
        if 재연설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(재연설비[a],i))
            if 재연설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(227,230):
        if 재연설비[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(재연설비[a],i))
            if 재연설비[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1

def 연결송수관클릭(load_wb):
    연결송수관 = ['N','N','N','N','N']

    load_ws = load_wb['13연결송수관']
    for i in range (1,6):
        연결송수관[i-1] = load_ws.cell(i,2).value

    a = 0
    for i in range(231,236):
        if 연결송수관[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(연결송수관[a],i))
            if 연결송수관[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1

def 비상콘센트_피난방화클릭(load_wb): #비상콘센트~피난방화
    비상콘센트_피난방화 = ['N','N','N','N','N','N','N','N','N']

    load_ws = load_wb['14비상콘센트_피난방화']
    for i in range (1,10):
        비상콘센트_피난방화[i-1] = load_ws.cell(i,2).value

    a = 0
    for i in range(237,241):
        if 비상콘센트_피난방화[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(비상콘센트_피난방화[a],i))
            if 비상콘센트_피난방화[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    if 비상콘센트_피난방화[a] != 'N':
        click('//*[@id="dtck{}_S0242"]'.format(비상콘센트_피난방화[a]))
        if 비상콘센트_피난방화[a] == 'B':
                click('//*[@id="itm_S0242SPL0001"]'.format(i))
    a+=1

    for i in range(244,248):
        if 비상콘센트_피난방화[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(비상콘센트_피난방화[a],i))
            if 비상콘센트_피난방화[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1

def 방염물품_화기취급클릭(load_wb):
    방염물품_화기취급 = [2,'N','N','N','N','N']

    load_ws = load_wb['17방염물품_화기취급']
    for i in range (1,7):
        방염물품_화기취급[i-1] = load_ws.cell(i,2).value

    a = 0
    click('//*[@id="itm_S0248SPL000{}"]'.format(방염물품_화기취급[a]))
    a+=1
    for i in range(249,251):
        if 방염물품_화기취급[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(방염물품_화기취급[a],i))
            if 방염물품_화기취급[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(252,255):
        if 방염물품_화기취급[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(방염물품_화기취급[a],i))
            if 방염물품_화기취급[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    
def 위험물저장클릭(load_wb):
    위험물저장 = ['N','N','N','N','N','N','N','N','N','N','N']

    load_ws = load_wb['19위험물저장']
    for i in range (1,12):
        위험물저장[i-1] = load_ws.cell(i,2).value

    a = 0
    for i in range(256,261):
        if 위험물저장[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(위험물저장[a],i))
            if 위험물저장[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1
    for i in range(262,273,2):
        if 위험물저장[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(위험물저장[a],i))
            if 위험물저장[a] == 'B':
                click('//*[@id="itm_S0{}SPL0001"]'.format(i))
        a+=1

def 소방활동관사클릭(load_wb):
    소방활동관사 = ['N','N','N','N','N','N','N']

    load_ws = load_wb['20소방활동관사']
    for i in range (1,8):
        소방활동관사[i-1] = load_ws.cell(i,2).value

    a = 0
    for i in range(274,278):
        if 소방활동관사[a] != 'N':
            click('//*[@id="dtck{}_S0{}"]'.format(소방활동관사[a],i))
            if 소방활동관사[a] == 'B':
                click('//*[@id="itm_S0{}SPL0002"]'.format(i))
        a+=1

    if 소방활동관사[a] != 'N':
            click('//*[@id="dtck{}_S0278"]'.format(소방활동관사[a]))
            if 소방활동관사[a] == 'B':
                click('//*[@id="itm_S0278SPL0001"]')
    a+=1

    if 소방활동관사[a] != 'N':
            click('//*[@id="dtck{}_S0279"]'.format(소방활동관사[a]))
            if 소방활동관사[a] == 'B':
                click('//*[@id="itm_S0279SPL0002"]')
    a+=1

    if 소방활동관사[a] != 'N':
            click('//*[@id="dtck{}_S0280"]'.format(소방활동관사[a]))
            if 소방활동관사[a] == 'B':
                click('//*[@id="itm_S0280SPL0001"]')

def 이용자특성클릭(load_wb):
    이용자특성 = [0,0,0,0,0,0,0]

    load_ws = load_wb['21이용자특성']
    for i in range (1,8):
        이용자특성[i-1] = load_ws.cell(i,2).value

    a = 0
    if 이용자특성[a] != 0:
        click('//*[@id="itm_S0282SPL000{}"]'.format(이용자특성[a]))
    a+=1
    for i in range(284,288):
        if 이용자특성[a] != 0:
            sendkey('//*[@id="dtck_S0{}"]'.format(i),이용자특성[a])
        a+=1
    for i in range(289,291):
        if 이용자특성[a] != 0:
            sendkey('//*[@id="dtck_S0{}"]'.format(i),이용자특성[a])
        a+=1

def 관리자특성클릭(load_wb): #관리자~끝까지
    관리자특성 = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,X,0,0,0,0,0]
    장애요인 = []
    load_ws = load_wb['22관리자특성']
    for i in range (1,21):
        관리자특성[i-1] = load_ws.cell(i,2).value
    for i in range(9,19):
        장애요인.append(load_ws.cell(i,6).value)

    a = 0
    for i in range(292,294): #2
        if 관리자특성[a] != 0:
            click('//*[@id="itm_S0{}SPL000{}"]'.format(i,관리자특성[a]))
        a+=1
    for i in range(295,299): #4
        if 관리자특성[a] != 0:
            click('//*[@id="itm_S0{}SPL000{}"]'.format(i,관리자특성[a]))
        a+=1
    for i in range(300,308): #8
        if 관리자특성[a] != 0:
            click('//*[@id="itm_S0{}SPL000{}"]'.format(i,관리자특성[a]))
        a+=1
    if 관리자특성[a] == 'x':
        click('//*[@id="itm_S0308SPL0011"]')
    elif 관리자특성[a] == 'o':
        for x in range(1,10):
            if 장애요인[x-1]=='o':
                click('//*[@id="itm_S0308SPL000{}"]'.format(x))
        if 장애요인[9]=='o':
            click('//*[@id="itm_S0308SPL0010"]')
    a+=1
    if 관리자특성[a] != 0:
        click('//*[@id="itm_S0309SPL000{}"]'.format(관리자특성[a]))
    a+=1
    for i in range(311,315): #4
        if 관리자특성[a] != 0:
            sendkey('//*[@id="dtck_S0{}"]'.format(i),관리자특성[a])
        a+=1     

def 저장_확인_클릭_로딩():
    click('//*[@id="b_Insert"]')
    time.sleep(0.5)
    click('//*[@id="btn_ok_c"]')
    time.sleep(0.5)
    click('//*[@id="btn_ok_a"]')
    time.sleep(0.5)
    check_loading()

def 건축_소방_입력():
    건축분야()
    저장_확인_클릭_로딩()

    소방분야()
    저장_확인_클릭_로딩()

def log_in():
    #로그인
    try:
        driver.find_element_by_id("_easyui_textbox_input4").click()
        driver.find_element_by_id("_easyui_combobox_i1_12").click()
        driver.find_element_by_id('_easyui_textbox_input1').send_keys('0105108374311')
        driver.find_element_by_id('_easyui_textbox_input2').send_keys('0000')
        time.sleep(0.5)
        driver.find_element_by_id("btnLogin").click()
        time.sleep(1)
    except:
        pass

def click(path):
    driver.find_element_by_xpath(path).click()

def sendkeyid(path,key):
    driver.find_element_by_id(path).send_keys(key)

def sendkey(path,key):
    driver.find_element_by_xpath(path).send_keys(key)

def 사이트_로그인_실행():
    global driver
    driver = webdriver.Chrome(chromedriver)
    driver.get('http://10.175.105.14/FSSE/lgn/login2.do')
    log_in()

switchnum = 0
def 탭전환1():
    global switchnum
    try:
        switchnum = switchnum * -1 + 1
        driver.switch_to_window(driver.window_handles[switchnum])
        실행중인탭.config(text='실행 탭 :'+ str(switchnum+1))
        
    except:
        pass

def 건물이름_get():
    global building_name
    building_name = 건물이름_입력.get()
    건물이름_현재.config(text=building_name)


def gui_program():
    global 실행중인탭
    global building_name
    global 건물이름_입력
    global 건물이름_현재
    root = Tk()
    root.title("특별조사 자동입력v2")
    root.geometry("350x400")

    building_name = '건물이름'

    row_num = 0

    사이트_로그인 = Button(root, width=5,height=1, text="Login", command=사이트_로그인_실행)
    사이트_로그인.grid(row=row_num, column=0, sticky=N+E+W+S, padx=3,pady=3)

    row_num+=1

    서명1버튼 = Button(root, width=16,height=3, text="서명1(이대진,송병진)", command=점검자서명1)
    서명1버튼.grid(row=row_num, column=1, sticky=N+E+W+S, padx=3,pady=3)
    서명2버튼 = Button(root, width=16,height=3, text="서명2(신학균,정조영)", command=점검자서명2)
    서명2버튼.grid(row=row_num, column=2, sticky=N+E+W+S, padx=3,pady=3)

    row_num+=1

    건물이름_입력 = Entry(root, width=20)
    건물이름_입력.grid(row=row_num, column=1, padx=3,pady=3)
    건물이름_입력.insert(0,"대상 폴더 입력")
    
    건물이름_현재 = Label(root,text=building_name)
    건물이름_현재.grid(row=row_num, column=2, sticky=N+E+W+S, padx=3,pady=3)
    
    row_num+=1
    
    건물이름_확인 = Button(root, width=3,height=1, text="확인", command=건물이름_get)
    건물이름_확인.grid(row=row_num, column=1, columnspan= 2, padx=3,pady=3)

    row_num+=1

    건축물개요버튼 = Button(root, width=32,height=3, text="건축물개요", command=건축물개요_입력1)
    건축물개요버튼.grid(row=row_num, column=1, columnspan= 2, sticky=N+E+W+S, padx=3,pady=3)
    
    row_num+=1

    건축버튼 = Button(root, width=16,height=3, text="건축", command=건축분야1)
    건축버튼.grid(row=row_num, column=1, sticky=N+E+W+S, padx=3,pady=3)
    소방버튼 = Button(root, width=16,height=3, text="소방", command=소방분야1)
    소방버튼.grid(row=row_num, column=2, sticky=N+E+W+S, padx=3,pady=3)
    
    row_num+=1

    건축_소방_저장 = Button(root, width=32,height=3, text="건축 소방 분야 완료", command=건축_소방_입력)
    건축_소방_저장.grid(row=row_num, column=1, columnspan= 2, sticky=N+E+W+S, padx=3,pady=3)

    row_num+=1

    탭전환 = Button(root, width=16,height=3, text="탭전환", command=탭전환1)
    탭전환.grid(row=row_num, column=1, sticky=N+E+W+S, padx=3,pady=3)

    실행중인탭 = Label(root,text='실행 탭 :'+ str(switchnum+1))
    실행중인탭.grid(row=row_num, column=2, sticky=N+E+W+S, padx=3,pady=3)


    root.mainloop()



if __name__ == '__main__':
    chromedriver = './chromedriver.exe'
    gui_program()
        
    



'''
    while True:
        x = input("입력하세요 : ") #건축물 개요 저장 후 엔터
        if x == '건축':
            try:
                건축분야()
                print("건축 fin")
            except:
                print("건축 error")
                pass
        elif x =='소방':
            try:
                소방분야()
                print("소방 fin")
            except:
                pass
        elif x == '서명1':
            try:
                점검자서명1()
                print("서명1 fin")
            except:
                pass
        elif x == '서명2':
            try:
                점검자서명2()
                print("서명2 fin")
            except:
                pass
        elif x == '건축물개요':

            try:
                건축물개요_입력()
                print("서명2 fin")
            except:
                pass
            
        else:
            print("명령없음")

'''