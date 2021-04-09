from json import load, loads
from re import search
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import *
import time

def get_excel(addr):
    load_wb = load_workbook('C:/Users/win10/Desktop/우편번호/주소목록1.xlsx', data_only=True)
    load_ws = load_wb['Sheet1']
    for i in range (1,500):
        addr.append(load_ws.cell(i,2).value)

def save_excel(name,address):
    wb = openpyxl.load_workbook('우편번호.xlsx')
    sheet1 = wb.active
    sheet1.append([name,address])
    wb.save('우편번호.xlsx')

def search_address(add):
    driver.find_element_by_xpath('//*[@id="nx_query"]').send_keys(Keys.CONTROL, 'a')
    driver.find_element_by_xpath('//*[@id="nx_query"]').send_keys(Keys.BACKSPACE)
    driver.find_element_by_xpath('//*[@id="nx_query"]').send_keys(add)
    driver.find_element_by_xpath('//*[@id="nx_search_form"]/fieldset/button/i').click()
    time.sleep(0.5)
    try:
        address_code = driver.find_element_by_xpath('//*[@id="loc-main-section-root"]/div/div/div[2]/div[2]/div/div[1]/div[2]/div/span[1]').text
        save_excel(add,address_code)
    except:
        save_excel(add,'error')
        pass

address = []
get_excel(address)
chromedriver = 'C:/Users/win10/Desktop/우편번호/chromedriver.exe'
driver = webdriver.Chrome(chromedriver)

driver.get('https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=1&ie=utf8&query=')

for x in range (1000):
    search_address(address[x])
    time.sleep(0.1)

time.sleep(1000)