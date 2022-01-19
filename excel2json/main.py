import sys
import openpyxl
import json

from PyQt5.QtWidgets import QApplication, QWidget, QMessageBox


class MyApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def Save(self):

        file_path = "자동화재속보설비 출동현황.xlsx"
        sheet_name = "비화재보 출동현황(24)"

        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet = wb[sheet_name]

        data = {}
        stat = {'By_Place': {'Dwelling_Place': {}, 'Etc_Place': {}, 'Factory_Place': {}, 'Senior_Place': {}},
                'Case_Stack': {
                    'Artificial_Factors': {'0': 0, '1': 0, '2': 0, '3': 0, '4': 0},
                    'Administrative_Factors': {'0': 0, '1': 0, '2': 0},
                    'System_Factors': {'0': 0, '1': 0, '2': 0, '3': 0},
                    'Etc_Factors': {'0': 0}
                }}

        # (2,7) ~ (30,~) 사용
        for row in sheet.iter_rows(min_row=7, max_row=sheet.max_row, min_col=2, max_col=30):
            # 3번째 열 (name)
            name = row[2].value

            # name이 false면 다음행으로 넘어감
            if not name:
                continue

            # data 딕셔너리에 name이 없으면 -> num = 1
            # ['Detail_Card'] 배열에 빈 딕셔너리 추가
            if name not in data:
                # num은 None 제외한 딕셔너리 개수
                num = 1
                data[name] = {}
                data[name]['Detail_Card'] = [None, ]
                data[name]['Detail_Card'].append({})
                
            # 있으면 -> num = ['Detail_Card'] 배열 길이
            # ['Detail_Card'] 배열에 빈 딕셔너리 추가
            else:
                num = len(data[name]['Detail_Card'])
                data[name]['Detail_Card'].append({})
            
            if None != row[1].value:
                data[name]['Declaration_Number_Phone'] = row[0].value
            else:
                data[name]['Declaration_Number_Phone'] = ""

            if None != row[3].value:
                data[name]['Old_Address'] = row[3].value
            else:
                data[name]['Old_Address'] = ""

            if None != row[4].value:
                data[name]['New_Address'] = row[4].value
            else:
                data[name]['New_Address'] = ""

            data[name]['Jurisdiction_Center'] = row[5].value

            if None != row[7].value:
                data[name]['Object_Manager'] = row[7].value
            else:
                data[name]['Object_Manager'] = ""

            if None != row[8].value:
                data[name]['Manager_General_Telephone'] = row[8].value
            else:
                data[name]['Manager_General_Telephone'] = ""

            if None != row[9].value:
                data[name]['Manager_Cell_Phone'] = row[9].value
            else:
                data[name]['Manager_Cell_Phone'] = ""

            data[name]['Object_Name'] = name
            data[name]['By_Place'] = ""
            data[name]['Num'] = str(len(data[name]['Detail_Card']) - 1)

            # index = data[name]['Detail_Card'][num]
            # data[name]['Detail_Card'][index] = {}
            days = str(row[1].value)
            years = days[0:4]
            month = days[4:6]

            if row[23].value == 1:
                data[name]['By_Place'] = '공장/창고'

                if years not in stat['By_Place']['Factory_Place']:
                    stat['By_Place']['Factory_Place'][years] = {}

                    if month not in stat['By_Place']['Factory_Place'][years]:
                        stat['By_Place']['Factory_Place'][years][month] = 1

                    else:
                        stat['By_Place']['Factory_Place'][years][month] += 1

                else:
                    if month not in stat['By_Place']['Factory_Place'][years]:
                        stat['By_Place']['Factory_Place'][years][month] = 1

                    else:
                        stat['By_Place']['Factory_Place'][years][month] += 1

            elif row[24].value == 1:
                data[name]['By_Place'] = '주거'

                if years not in stat['By_Place']['Dwelling_Place']:
                    stat['By_Place']['Dwelling_Place'][years] = {}
                    if month not in stat['By_Place']['Dwelling_Place'][years]:
                        stat['By_Place']['Dwelling_Place'][years][month] = 1

                    else:
                        stat['By_Place']['Dwelling_Place'][years][month] += 1

                else:
                    if month not in stat['By_Place']['Dwelling_Place'][years]:
                        stat['By_Place']['Dwelling_Place'][years][month] = 1

                    else:
                        stat['By_Place']['Dwelling_Place'][years][month] += 1

            elif row[25].value == 1:
                data[name]['By_Place'] = '노유자'

                if years not in stat['By_Place']['Senior_Place']:
                    stat['By_Place']['Senior_Place'][years] = {}
                    if month not in stat['By_Place']['Senior_Place'][years]:
                        stat['By_Place']['Senior_Place'][years][month] = 1

                    else:
                        stat['By_Place']['Senior_Place'][years][month] += 1

                else:
                    if month not in stat['By_Place']['Senior_Place'][years]:
                        stat['By_Place']['Senior_Place'][years][month] = 1

                    else:
                        stat['By_Place']['Senior_Place'][years][month] += 1

            elif row[26].value == 1:
                data[name]['By_Place'] = '기타'

                if years not in stat['By_Place']['Etc_Place']:
                    stat['By_Place']['Etc_Place'][years] = {}
                    if month not in stat['By_Place']['Etc_Place'][years]:
                        stat['By_Place']['Etc_Place'][years][month] = 1

                    else:
                        stat['By_Place']['Etc_Place'][years][month] += 1

                else:
                    if month not in stat['By_Place']['Etc_Place'][years]:
                        stat['By_Place']['Etc_Place'][years][month] = 1

                    else:
                        stat['By_Place']['Etc_Place'][years][month] += 1

            data[name]['Detail_Card'][num]['Reporting_Time'] = days[0:4] + "년 " + days[4:6] + "월 " + days[
                                                                                                     6:8] + "일 " + days[8:10] + "시 " + days[10:12] + "분"

            data[name]['Detail_Card'][num]['Reported_Content'] = row[6].value
            data[name]['Detail_Card'][num]['By_Case_Cause'] = ""
            data[name]['Detail_Card'][num]['Factors_Position'] = ""
            data[name]['Detail_Card'][num]['Factors_Stack'] = ""
            data[name]['Detail_Card'][num]['Object_Name'] = name

            for i in range(10, 23):
                if row[i].value == 1:
                    #data[name]['Detail_Card'][num]['By_Case_Cause'] = sheet.cell(5, i + 3).value
                    data[name]['Detail_Card'][num]['By_Case_Cause'] = sheet.cell(5, i + 2).value
                    if i <= 14:
                        data[name]['Detail_Card'][num]['Factors_Stack'] = 'Artificial_Factors'
                        stat['Case_Stack']['Artificial_Factors'][str(i - 10)] += 1
                        data[name]['Detail_Card'][num]['Factors_Position'] = str(i - 10)
                    elif i <= 17:
                        data[name]['Detail_Card'][num]['Factors_Stack'] = 'Administrative_Factors'
                        stat['Case_Stack']['Administrative_Factors'][str(i - 15)] += 1
                        data[name]['Detail_Card'][num]['Factors_Position'] = str(i - 15)
                    elif i <= 21:
                        data[name]['Detail_Card'][num]['Factors_Stack'] = 'System_Factors'
                        stat['Case_Stack']['System_Factors'][str(i - 18)] += 1
                        data[name]['Detail_Card'][num]['Factors_Position'] = str(i - 18)
                    else:
                        data[name]['Detail_Card'][num]['Factors_Stack'] = 'Etc_Factors'
                        stat['Case_Stack']['Etc_Factors'][str(i - 22)] += 1
                        data[name]['Detail_Card'][num]['Factors_Position'] = str(i - 22)
                    break

        result = {'Data': data, 'Statistics': stat}

        with open('data.json', 'w', encoding='UTF8') as f:
            json.dump(result, f, ensure_ascii=False, indent=4)

    def initUI(self):
        self.Save()
        QMessageBox.about(self, '불러오기 성공', '파일 불러오기 성공')


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())
